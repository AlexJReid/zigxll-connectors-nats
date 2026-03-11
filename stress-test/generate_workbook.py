#!/usr/bin/env python3
"""Generate a stress-test .xlsx with many =NATS.SUB() RTD formulas.

Usage:
    python generate_workbook.py [--subjects N] [--cells N] [--output FILE]

Defaults: 200 unique subjects, 500 total cells (so ~60% are duplicates).
Subjects follow the pattern: stress.{category}.{id}
"""

import argparse
import random
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

CATEGORIES = [
    "prices", "sensors", "metrics", "orders", "events",
    "ticks", "status", "alerts", "logs", "telemetry",
]


def generate_subjects(n: int) -> list[str]:
    """Generate n unique NATS subjects across categories."""
    subjects = []
    per_cat = max(1, n // len(CATEGORIES))
    for cat in CATEGORIES:
        for i in range(per_cat):
            subjects.append(f"stress.{cat}.{i}")
            if len(subjects) >= n:
                return subjects
    # fill remainder
    while len(subjects) < n:
        cat = random.choice(CATEGORIES)
        subjects.append(f"stress.{cat}.{len(subjects)}")
    return subjects


def build_workbook(subjects: list[str], total_cells: int, output: str):
    wb = Workbook()

    # --- Sheet 1: Main stress grid ---
    ws = wb.active
    ws.title = "Stress Test"

    # Pick cells with duplicates
    cells = [random.choice(subjects) for _ in range(total_cells)]

    # Header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    cols_per_row = 10
    for c in range(1, cols_per_row + 1):
        cell = ws.cell(row=1, column=c, value=f"Sub {c}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = 22

    # Fill formulas
    row = 2
    col = 1
    for subj in cells:
        ws.cell(row=row, column=col, value=f'=NATS.SUB("{subj}")')
        col += 1
        if col > cols_per_row:
            col = 1
            row += 1

    # --- Sheet 2: Summary / legend ---
    ws2 = wb.create_sheet("Info")
    ws2["A1"] = "Stress Test Workbook"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Unique subjects:"
    ws2["B3"] = len(subjects)
    ws2["A4"] = "Total RTD cells:"
    ws2["B4"] = total_cells
    ws2["A5"] = "Duplicate ratio:"
    ws2["B5"] = f"{(1 - len(subjects) / total_cells) * 100:.0f}%"
    ws2["A7"] = "Subject pattern:"
    ws2["B7"] = "stress.{category}.{id}"
    ws2["A8"] = "Categories:"
    ws2["B8"] = ", ".join(CATEGORIES)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 50

    # --- Sheet 3: Per-category breakdown ---
    ws3 = wb.create_sheet("By Category")
    ws3["A1"] = "Category"
    ws3["B1"] = "Subject"
    ws3["C1"] = "Value"
    for c in ["A", "B", "C"]:
        ws3[f"{c}1"].font = Font(bold=True)
    ws3.column_dimensions["A"].width = 15
    ws3.column_dimensions["B"].width = 30
    ws3.column_dimensions["C"].width = 22

    row = 2
    for subj in sorted(subjects):
        cat = subj.split(".")[1]
        ws3.cell(row=row, column=1, value=cat)
        ws3.cell(row=row, column=2, value=subj)
        ws3.cell(row=row, column=3, value=f'=NATS.SUB("{subj}")')
        row += 1

    wb.save(output)
    print(f"Wrote {output}")
    print(f"  {len(subjects)} unique subjects, {total_cells} total RTD cells")
    print(f"  Sheets: 'Stress Test' (grid), 'Info' (summary), 'By Category' (sorted)")


def main():
    parser = argparse.ArgumentParser(description="Generate NATS RTD stress-test workbook")
    parser.add_argument("--subjects", type=int, default=200, help="Number of unique subjects (default: 200)")
    parser.add_argument("--cells", type=int, default=500, help="Total RTD formula cells (default: 500)")
    parser.add_argument("--output", default="stress-test.xlsx", help="Output filename (default: stress-test.xlsx)")
    args = parser.parse_args()

    if args.cells < args.subjects:
        parser.error("--cells must be >= --subjects to fit all unique subjects")

    subjects = generate_subjects(args.subjects)
    build_workbook(subjects, args.cells, args.output)


if __name__ == "__main__":
    main()
